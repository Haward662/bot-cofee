import logging
from telegram import (
    Update,
    ReplyKeyboardMarkup,
    KeyboardButton,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    InputMediaPhoto,
)
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    ConversationHandler,
    ContextTypes,
    filters,
    CallbackQueryHandler,
)
import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os

# Настройка логирования
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO
)
logger = logging.getLogger(__name__)

# Токен бота
BOT_TOKEN = "8432720196:AAF1Tzf9R8Ehv1QflO0Dc121clDJb7noXC8"

# ID администраторов (добавьте свои ID через запятую)
# Чтобы узнать свой ID, напишите боту @userinfobot
ADMIN_IDS = []  # Пример: [123456789, 987654321]

# Состояния для ConversationHandler
WAITING_EMAIL = 1

# Имя файла Excel
EXCEL_FILE = "bot_data.xlsx"

# Категории товаров с описаниями (для карусели)
PRODUCT_CATEGORIES = [
    {
        "id": "cafa_france",
        "name": "Очки воителя Cafa France",
        "description": "Стильные и надежные очки для водителей",
        "images": [
            "1_ЛКЛ_Выкладка_на_стойке_ов_в_бордовом_оформлении_очки.jpg"
        ],  # Пути к изображениям
    },
    {
        "id": "driver",
        "name": "Аксессуары Driver",
        "description": "Качественные аксессуары для автомобиля",
        "images": [
            "5_Олви_Выкладка_хд+ножи_шф_на_стойке_60_см_4.jpg",
            "5 Драйвер_2.jpg",
        ],
    },
    {
        "id": "takeshi",
        "name": "Ножи и фонари Takeshi",
        "description": "Профессиональные ножи и фонари",
        "images": ["3_Тебойл_Выкладка_Такеши_ножи_фонари_на_стойке_60_см_2025.jpg"],
    },
    {
        "id": "chef_ferguson",
        "name": "Ножи и товары для кухни Chef Ferguson",
        "description": "Кухонные принадлежности высокого качества",
        "images": ["4_Татнефть_выкладка_кухня_на_стойке_2025.jpg"],
    },
    {
        "id": "norge",
        "name": "Зимние товары NORGE",
        "description": "Товары для зимнего сезона",
        "images": ["2_Даль_Норге_2025_выкладка_на_стойке_60_см.jpg"],
    },
]


# Инициализация базы данных
def init_db():
    conn = sqlite3.connect("bot_database.db")
    cursor = conn.cursor()

    # Таблица пользователей
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            user_id INTEGER PRIMARY KEY,
            username TEXT,
            first_name TEXT,
            last_name TEXT,
            phone_number TEXT,
            email TEXT,
            registered_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """
    )

    # Таблица выданных промокодов
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS coffee_issued (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            issued_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users (user_id)
        )
    """
    )

    conn.commit()
    conn.close()
    # Инициализация Excel файла
    init_excel()


# Функции для работы с Excel
def init_excel():
    """Создает Excel файл с заголовками, если его нет"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()

        # Лист пользователей
        ws_users = wb.active
        ws_users.title = "Пользователи"
        headers_users = [
            "ID пользователя",
            "Username",
            "Имя",
            "Фамилия",
            "Email",
            "Дата регистрации",
        ]
        ws_users.append(headers_users)

        # Стилизация заголовков
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws_users[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Лист выданных кофе
        ws_coffee = wb.create_sheet("Выданные кофе")
        headers_coffee = ["ID", "ID пользователя", "Username", "Имя", "Дата выдачи"]
        ws_coffee.append(headers_coffee)

        for cell in ws_coffee[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Настройка ширины колонок
        ws_users.column_dimensions["A"].width = 15
        ws_users.column_dimensions["B"].width = 20
        ws_users.column_dimensions["C"].width = 20
        ws_users.column_dimensions["D"].width = 20
        ws_users.column_dimensions["E"].width = 30
        ws_users.column_dimensions["F"].width = 20

        ws_coffee.column_dimensions["A"].width = 10
        ws_coffee.column_dimensions["B"].width = 15
        ws_coffee.column_dimensions["C"].width = 20
        ws_coffee.column_dimensions["D"].width = 20
        ws_coffee.column_dimensions["E"].width = 20

        wb.save(EXCEL_FILE)
        logger.info(f"Excel файл {EXCEL_FILE} создан")


def save_user_to_excel(
    user_id, username=None, first_name=None, last_name=None, email=None
):
    """Сохраняет пользователя в Excel"""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Пользователи"]

        # Проверяем, есть ли уже такой пользователь
        user_exists = False
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[0].value == user_id:
                user_exists = True
                # Обновляем данные, если они изменились
                if username and row[1].value != username:
                    row[1].value = username
                if first_name and row[2].value != first_name:
                    row[2].value = first_name
                if last_name and row[3].value != last_name:
                    row[3].value = last_name
                if email and row[4].value != email:
                    row[4].value = email
                break

        if not user_exists:
            # Добавляем нового пользователя
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append(
                [
                    user_id,
                    username or "",
                    first_name or "",
                    last_name or "",
                    email or "",
                    now,
                ]
            )

        wb.save(EXCEL_FILE)
    except Exception as e:
        logger.error(f"Ошибка при сохранении в Excel: {e}")


def save_coffee_to_excel(coffee_id, user_id, username=None, first_name=None):
    """Сохраняет выдачу кофе в Excel"""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Выданные кофе"]

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([coffee_id, user_id, username or "", first_name or "", now])

        wb.save(EXCEL_FILE)
    except Exception as e:
        logger.error(f"Ошибка при сохранении кофе в Excel: {e}")


# Функции для работы с БД
def save_user(user_id, username=None, first_name=None, last_name=None):
    conn = sqlite3.connect("bot_database.db")
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT OR IGNORE INTO users (user_id, username, first_name, last_name)
        VALUES (?, ?, ?, ?)
    """,
        (user_id, username, first_name, last_name),
    )
    conn.commit()
    conn.close()
    # Сохраняем в Excel
    save_user_to_excel(user_id, username, first_name, last_name)


def save_email(user_id, email):
    conn = sqlite3.connect("bot_database.db")
    cursor = conn.cursor()
    cursor.execute(
        """
        UPDATE users SET email = ? WHERE user_id = ?
    """,
        (email, user_id),
    )
    # Получаем данные пользователя для Excel
    cursor.execute(
        "SELECT username, first_name, last_name FROM users WHERE user_id = ?",
        (user_id,),
    )
    user_data = cursor.fetchone()
    conn.commit()
    conn.close()
    # Обновляем в Excel
    if user_data:
        save_user_to_excel(user_id, user_data[0], user_data[1], user_data[2], email)


def issue_coffee(user_id):
    conn = sqlite3.connect("bot_database.db")
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO coffee_issued (user_id) VALUES (?)
    """,
        (user_id,),
    )
    coffee_id = cursor.lastrowid
    conn.commit()

    # Получаем данные пользователя для Excel
    cursor.execute(
        "SELECT username, first_name FROM users WHERE user_id = ?", (user_id,)
    )
    user_data = cursor.fetchone()
    username = user_data[0] if user_data else None
    first_name = user_data[1] if user_data else None

    conn.close()
    # Сохраняем в Excel
    save_coffee_to_excel(coffee_id, user_id, username, first_name)


def get_statistics():
    conn = sqlite3.connect("bot_database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM users")
    total_users = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM coffee_issued")
    total_coffee = cursor.fetchone()[0]
    conn.close()
    return total_users, total_coffee


# Создание клавиатуры главного меню
def get_main_keyboard():
    keyboard = [
        [KeyboardButton("Получить кофе бесплатно")],
        [KeyboardButton("Посмотреть, за покупку каких товаров дают кофе бесплатно")],
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)


# Создание клавиатуры для подтверждения покупки
def get_purchase_confirmation_keyboard():
    keyboard = [[KeyboardButton("Да")], [KeyboardButton("Нет")]]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)


# Создание клавиатуры после выдачи кофе
def get_after_coffee_keyboard():
    keyboard = [[KeyboardButton("Вернуться к началу")]]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)


# Путь к баннеру (используем абсолютный путь относительно файла bot.py)
BANNER_PHOTO = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "photo_2025-12-15_17-00-19.jpg"
)


def get_cafafrancebot_keyboard() -> InlineKeyboardMarkup:
    """Клавиатура с кнопкой перехода во второй бот @cafafrancebot."""
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(
                    "Выгода до 5000 руб. в @cafafrancebot",
                    url="https://t.me/cafafrancebot",
                )
            ]
        ]
    )


# Функция для отправки рекламного баннера второго бота
async def send_banner_photo(message):
    """
    Отправляет баннер-фото и кнопку с ссылкой на @cafafrancebot
    последним сообщением.
    """
    if message is None:
        logger.error("Message объект равен None, невозможно отправить баннер")
        return

    caption = "Вместе с кофе забирай выгоду до 5000 руб. в боте @cafafrancebot"

    # Если есть файл баннера — отправляем как фото с подписью и кнопкой
    if os.path.exists(BANNER_PHOTO):
        try:
            logger.info(f"Попытка отправить баннер: {BANNER_PHOTO}")
            with open(BANNER_PHOTO, "rb") as photo:
                await message.reply_photo(
                    photo=photo,
                    caption=caption,
                    reply_markup=get_cafafrancebot_keyboard(),
                )
            logger.info(f"Баннер успешно отправлен: {BANNER_PHOTO}")
            return
        except Exception as e:
            logger.error(f"Ошибка при отправке баннера: {e}", exc_info=True)

    # Если файл не найден или произошла ошибка — шлём просто текст + кнопку
    logger.warning(
        f"Баннер-файл не отправлен, используем текстовый блок для @cafafrancebot "
        f"(путь: {BANNER_PHOTO}, cwd: {os.getcwd()})"
    )
    try:
        await message.reply_text(
            caption,
            reply_markup=get_cafafrancebot_keyboard(),
        )
    except Exception as e:
        logger.error(f"Ошибка при отправке текстового баннера: {e}", exc_info=True)


# Обработчик команды /start
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    save_user(user.id, user.username, user.first_name, user.last_name)

    welcome_text = (
        "☕ Добро пожаловать в акцию «Получи кофе бесплатно»!\n\n" "Выберите действие:"
    )

    await update.message.reply_text(welcome_text, reply_markup=get_main_keyboard())


# Функция создания клавиатуры карусели
def get_carousel_keyboard(category_index):
    """Создает клавиатуру для карусели категорий"""
    total = len(PRODUCT_CATEGORIES)
    prev_index = (category_index - 1) % total
    next_index = (category_index + 1) % total

    keyboard = []

    # Кнопки навигации
    nav_buttons = []
    if total > 1:
        nav_buttons.append(
            InlineKeyboardButton("◀️ Назад", callback_data=f"category_{prev_index}")
        )
        nav_buttons.append(
            InlineKeyboardButton(
                f"{category_index + 1}/{total}", callback_data="category_info"
            )
        )
        nav_buttons.append(
            InlineKeyboardButton("Вперед ▶️", callback_data=f"category_{next_index}")
        )
        keyboard.append(nav_buttons)

    # Кнопки действий
    keyboard.append(
        [InlineKeyboardButton("Понятно, хочу кофе", callback_data="want_coffee")]
    )
    keyboard.append(
        [InlineKeyboardButton("Вернуться к началу", callback_data="back_to_start")]
    )

    return InlineKeyboardMarkup(keyboard)


# Обработчик просмотра товаров (показывает все категории сразу)
async def show_products(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает все категории подряд, чтобы пользователь мог промотать."""
    # Отправляем все категории одну за другой
    for category in PRODUCT_CATEGORIES:
        products_text = (
            f"🛍️ **{category['name']}**\n\n"
            f"{category['description']}\n\n"
            f"💡 **Условия акции:**\n"
            f"Собери покупку на сумму от 1 900 руб. товарами этих брендов — и получи кофе в подарок."
        )

        # Берём первое изображение категории (если есть)
        image_path = None
        if category.get("images"):
            for p in category["images"]:
                if os.path.exists(p):
                    image_path = p
                    break

        if image_path:
            with open(image_path, "rb") as photo:
                await update.message.reply_photo(
                    photo=photo,
                    caption=products_text,
                    parse_mode="Markdown",
                )
        else:
            await update.message.reply_text(
                products_text,
                parse_mode="Markdown",
            )

    # В конце отправляем кнопки
    keyboard = [
        [KeyboardButton("Понятно, хочу кофе")],
        [KeyboardButton("Вернуться к началу")],
    ]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

    await update.message.reply_text("Выберите действие:", reply_markup=reply_markup)

    # Отправляем баннер последним сообщением
    await send_banner_photo(update.message)


# Обработчик callback для карусели
async def handle_category_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data == "back_to_start":
        # Удаляем сообщение с каруселью и отправляем главное меню
        await query.message.delete()
        welcome_text = (
            "☕ Добро пожаловать в акцию «Получи кофе бесплатно»!\n\n"
            "Выберите действие:"
        )
        await query.message.reply_text(welcome_text, reply_markup=get_main_keyboard())
        return

    if query.data == "want_coffee":
        # Удаляем сообщение с каруселью и переходим к запросу кофе
        await query.message.delete()
        question_text = (
            "❓ Вы приобрели товары участвующих брендов на сумму от 1 900 руб.?"
        )
        await query.message.reply_text(
            question_text, reply_markup=get_purchase_confirmation_keyboard()
        )
        return

    if query.data.startswith("category_"):
        try:
            category_index = int(query.data.split("_")[1])
            category = PRODUCT_CATEGORIES[category_index]

            products_text = (
                f"🛍️ **{category['name']}**\n\n"
                f"{category['description']}\n\n"
                f"💡 **Условия акции:**\n"
                f"Собери покупку на сумму от 1 900 руб. товарами этих брендов — и получи кофе в подарок."
            )

            reply_markup = get_carousel_keyboard(category_index)

            # Берём первое доступное изображение, если оно есть
            image_path = None
            if category.get("images"):
                for p in category["images"]:
                    if os.path.exists(p):
                        image_path = p
                        break

            if image_path:
                # Меняем картинку и подпись в том же сообщении (замена поста)
                with open(image_path, "rb") as photo:
                    await query.edit_message_media(
                        media=InputMediaPhoto(
                            media=photo,
                            caption=products_text,
                            parse_mode="Markdown",
                        ),
                        reply_markup=reply_markup,
                    )
            else:
                # Меняем только текст и кнопки
                await query.edit_message_text(
                    products_text,
                    reply_markup=reply_markup,
                    parse_mode="Markdown",
                )
        except (ValueError, IndexError):
            await query.answer("Ошибка при переключении категории", show_alert=True)
        except Exception as e:
            logger.error(f"Ошибка при отправке категории: {e}")
            await query.answer("Ошибка при загрузке изображений", show_alert=True)


# Обработчик запроса на получение кофе
async def request_coffee(update: Update, context: ContextTypes.DEFAULT_TYPE):
    question_text = "❓ Вы приобрели товары участвующих брендов на сумму от 1 900 руб.?"

    await update.message.reply_text(
        question_text, reply_markup=get_purchase_confirmation_keyboard()
    )


# Обработчик подтверждения покупки (Да)
async def confirm_purchase(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    issue_coffee(user_id)

    coffee_text = "✅ Отлично! Обратитесь к оператору и получите бесплатный кофе."

    keyboard = [
        [KeyboardButton("Оставить email")],
        [KeyboardButton("Вернуться к началу")],
    ]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

    await update.message.reply_text(coffee_text, reply_markup=reply_markup)

    # Отправляем баннер последним сообщением
    await send_banner_photo(update.message)


# Обработчик отказа от покупки (Нет)
async def deny_purchase(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await show_products(update, context)


# Обработчик запроса email
async def request_email(update: Update, context: ContextTypes.DEFAULT_TYPE):
    email_text = (
        "📧 Хотите получать другие подарки и акции? Оставьте ваш email.\n\n"
        "Если не хотите делиться email — просто нажмите «Вернуться к началу» 👇"
    )

    await update.message.reply_text(email_text)
    return WAITING_EMAIL


# Обработчик ввода email
async def receive_email(update: Update, context: ContextTypes.DEFAULT_TYPE):
    email = update.message.text

    # Если пользователь нажал любую из основных кнопок — выходим из ввода email
    navigation_texts = {
        "Вернуться к началу",
        "Получить кофе бесплатно",
        "Понятно, хочу кофе",
        "Посмотреть, за покупку каких товаров дают кофе бесплатно",
        "Да",
        "Нет",
    }
    if email in navigation_texts:
        # Завершаем состояние и передаем сообщение в общий обработчик
        await handle_text(update, context)
        return ConversationHandler.END

    # Простая валидация email
    if "@" in email and "." in email:
        user_id = update.effective_user.id
        save_email(user_id, email)

        thank_you_text = (
            "🙏 Спасибо! Мы будем присылать вам информацию о новых акциях и подарках."
        )

        await update.message.reply_text(
            thank_you_text, reply_markup=get_after_coffee_keyboard()
        )
        # Отправляем баннер последним сообщением
        await send_banner_photo(update.message)
        return ConversationHandler.END
    else:
        # Не ругаемся, а мягко даём шанс попробовать ещё раз или вернуться
        await update.message.reply_text(
            "❌ Пожалуйста, введите корректный email адрес "
            "или нажмите «Вернуться к началу», чтобы пропустить этот шаг."
        )
        return WAITING_EMAIL


# Обработчик возврата к началу
async def return_to_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await start(update, context)
    return ConversationHandler.END


# Обработчик текстовых сообщений
async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text

    if text == "Получить кофе бесплатно":
        await request_coffee(update, context)
    elif text == "Посмотреть, за покупку каких товаров дают кофе бесплатно":
        await show_products(update, context)
    elif text == "Понятно, хочу кофе":
        await request_coffee(update, context)
    elif text == "Да":
        await confirm_purchase(update, context)
    elif text == "Нет":
        await deny_purchase(update, context)
    elif text == "Вернуться к началу":
        await return_to_start(update, context)
    else:
        await update.message.reply_text(
            "Пожалуйста, используйте кнопки для навигации.",
            reply_markup=get_main_keyboard(),
        )


# Обработчик команды /stats (для администратора)
async def stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    total_users, total_coffee = get_statistics()
    stats_text = (
        f"📊 **Статистика бота:**\n\n"
        f"👥 Всего пользователей: {total_users}\n"
        f"☕ Выдано кофе: {total_coffee}"
    )
    await update.message.reply_text(stats_text, parse_mode="Markdown")


# Обработчик команды /export (для администратора) - отправка Excel файла
async def export_data(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id

    # Проверка прав администратора (если список пустой, доступ открыт для всех)
    if ADMIN_IDS and user_id not in ADMIN_IDS:
        await update.message.reply_text(
            "❌ У вас нет прав для выполнения этой команды."
        )
        return

    if not os.path.exists(EXCEL_FILE):
        await update.message.reply_text("❌ Excel файл еще не создан. Данных пока нет.")
        return

    try:
        await update.message.reply_document(
            document=open(EXCEL_FILE, "rb"),
            filename=EXCEL_FILE,
            caption="📊 Экспорт данных бота",
        )
        logger.info(f"Администратор {user_id} экспортировал данные")
    except Exception as e:
        logger.error(f"Ошибка при экспорте данных: {e}")
        await update.message.reply_text(f"❌ Ошибка при отправке файла: {e}")


# Отмена ввода email
async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Отменено.", reply_markup=get_main_keyboard())
    return ConversationHandler.END


def main():
    # Инициализация БД
    init_db()

    # Создание приложения
    application = Application.builder().token(BOT_TOKEN).build()

    # ConversationHandler для сбора email
    email_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^Оставить email$"), request_email)],
        states={
            WAITING_EMAIL: [
                # Сначала даём приоритет кнопке «Вернуться к началу»
                MessageHandler(filters.Regex("^Вернуться к началу$"), return_to_start),
                # Всё остальное считаем попыткой ввода email
                MessageHandler(filters.TEXT & ~filters.COMMAND, receive_email),
            ]
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    # Регистрация обработчиков (важен порядок - ConversationHandler должен быть перед общим MessageHandler)
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("stats", stats))
    application.add_handler(CommandHandler("export", export_data))
    application.add_handler(CallbackQueryHandler(handle_category_callback))
    application.add_handler(email_handler)
    application.add_handler(
        MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text)
    )

    # Запуск бота
    logger.info("Бот запущен...")
    application.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
